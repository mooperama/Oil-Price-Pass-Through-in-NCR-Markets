# """
# utils.py
# ========
# Requires: pdfplumber, pandas, openpyxl, requests, beautifulsoup4, python-dotenv
# Install : pip install pdfplumber pandas openpyxl requests beautifulsoup4 python-dotenv
# """

# import os
# import re
# import requests
# import pdfplumber
# import itertools
# import pandas as pd
# from sqlalchemy import create_engine
# from pathlib import Path
# from bs4 import BeautifulSoup
# from urllib.parse import urljoin
# from dotenv import load_dotenv
# from datetime import datetime
# from openpyxl import Workbook
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.utils import get_column_letter

# # Load environment variables from the .env file
# load_dotenv()

# # ==========================================
# # 00.1 Data Acquisition EIA Brent Prices
# # ==========================================
# def fetch_eia_prices(series_id, price_name):
#     """Fetches daily spot prices for a given EIA series ID."""
#     api_key = os.getenv('EIA_API_KEY')
#     start_date = os.getenv('START_DATE')

#     print(f"Fetching {price_name} prices from {start_date}...")

#     url = (
#         f"https://api.eia.gov/v2/petroleum/pri/spt/data/"
#         f"?api_key={api_key}"
#         f"&frequency=daily"
#         f"&data[0]=value"
#         f"&facets[series][]={series_id}"
#         f"&start={start_date}"
#         f"&sort[0][column]=period"
#         f"&sort[0][direction]=desc"
#         f"&length=5000"
#     )

#     response = requests.get(url)

#     if response.status_code == 200:
#         data = response.json()['response']['data']
#         df = pd.DataFrame(data)

#         # Clean and isolate the columns
#         df = df[['period', 'value']]
#         df.rename(columns={'period': 'Date', 'value': price_name}, inplace=True)

#         # Convert to datetime and strip timezone
#         df['Date'] = pd.to_datetime(df['Date'])
#         if df['Date'].dt.tz is not None:
#             df['Date'] = df['Date'].dt.tz_localize(None)

#         return df
#     else:
#         print(f"Failed to fetch {price_name}. Status: {response.status_code}")
#         return pd.DataFrame()


# def download_eia_dataset(output_folder='data'):
#     """Executes the pipeline to pull and merge Brent and WTI prices."""
#     if not os.path.exists(output_folder):
#         os.makedirs(output_folder)

#     df_brent = fetch_eia_prices('RBRTE', 'Brent_Price_USD')
#     df_wti = fetch_eia_prices('RWTC', 'WTI_Price_USD')

#     if not df_brent.empty and not df_wti.empty:
#         print("Merging Brent and WTI datasets...")
#         df_combined = pd.merge(df_brent, df_wti, on='Date', how='inner')
#         df_combined.sort_values(by='Date', inplace=True)

#         output_file = os.path.join(output_folder, 'Brent_WTI_Prices_2020_to_Present.csv')
#         df_combined.to_csv(output_file, index=False)
#         print(f"Success! Saved EIA data to {output_file}")
#     else:
#         print("Error: Could not complete merge due to missing data.")


# # ==========================================
# # 00.2 Data Acquisition DOE
# # ==========================================
# def download_doe_pdfs(output_folder='data'):
#     """Scrapes the DOE portal for retail pump price PDFs."""
#     target_url = os.getenv('DOE_TARGET_URL')
#     session_cookie = os.getenv('DOE_SESSION_COOKIE')

#     if not os.path.exists(output_folder):
#         os.makedirs(output_folder)

#     headers = {
#         'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0 Safari/537.36',
#         'Cookie': session_cookie
#     }

#     print("Authenticating with DOE portal...")
#     response = requests.get(target_url, headers=headers)

#     if "Log in" in response.text or response.status_code != 200:
#         print("Recheck Cookie Validity. Authentication failed.")
#         return

#     soup = BeautifulSoup(response.text, 'html.parser')
#     pdf_links = []

#     for link in soup.find_all('a', href=True):
#         href = link['href']
#         if 'documents/d/' in href.lower() or '-pdf' in href.lower() or href.lower().endswith('.pdf'):
#             full_url = urljoin(target_url, href)
#             pdf_links.append(full_url)

#     pdf_links = list(set(pdf_links))
#     print(f"Found {len(pdf_links)} protected PDF(s). Downloading...")

#     for pdf_url in pdf_links:
#         file_name = pdf_url.split('/')[-1]

#         if not file_name.endswith('.pdf'):
#             file_name = file_name + '.pdf'

#         file_path = os.path.join(output_folder, file_name)

#         pdf_response = requests.get(pdf_url, headers=headers)
#         pdf_response.raise_for_status()

#         with open(file_path, 'wb') as file:
#             file.write(pdf_response.content)
#         print(f"Saved: {file_name}")

#     print("DOE PDF extraction complete.")


# # ==========================================
# # 1.1 Preprocessing DOE filename
# # ==========================================
# def standardize_doe_date(data_folder='data/doe_pump_prices'):
#     """
#     Loops through the target folder, extracts dates from messy DOE filenames,
#     and physically renames the PDF files to 'YYYY-MM-DD.pdf' on the OS.
#     """
#     renamed_count = 0

#     for filename in os.listdir(data_folder):
#         if filename.lower().endswith('.pdf'):

#             year, month, day = None, None, None

#             # Pattern 1: YYYY-MM-DD or YYYY-Month-DD (e.g., 2021-june-10, 2024-dec-24)
#             m1 = re.search(r"(20\d{2})[-_]([a-z]{3,9}|\d{1,2})[-_](\d{1,2})", filename.lower())

#             # Pattern 2: MM-DD-YYYY (e.g., 07-01-2025)
#             m2 = re.search(r"(\d{1,2})[-_](\d{1,2})[-_](20\d{2})", filename.lower())

#             # Pattern 3: MMDDYYYY contiguous (e.g., 08122025)
#             m3 = re.search(r"(\d{2})(\d{2})(20\d{2})", filename.lower())

#             # Check which pattern matched
#             if m1:
#                 year, month, day = m1.groups()
#             elif m2:
#                 month, day, year = m2.groups()
#             elif m3:
#                 month, day, year = m3.groups()
#             else:
#                 print(f"  -> WARNING: Could not find any date pattern in '{filename}'")
#                 continue

#             day = day.zfill(2)
#             month = str(month)

#             try:
#                 # Handle textual months (forcing them to 3 letters: June -> Jun, Sept -> Sep)
#                 if month.isalpha():
#                     month_short = month[:3].title()
#                     clean_str = f"{year}-{month_short}-{day}"
#                     clean_date = datetime.strptime(clean_str, "%Y-%b-%d").strftime("%Y-%m-%d")
#                 # Handle numeric months
#                 else:
#                     month = month.zfill(2)
#                     clean_str = f"{year}-{month}-{day}"
#                     clean_date = datetime.strptime(clean_str, "%Y-%m-%d").strftime("%Y-%m-%d")

#                 # Construct the new file path
#                 new_filename = f"{clean_date}.pdf"
#                 old_path = os.path.join(data_folder, filename)
#                 new_path = os.path.join(data_folder, new_filename)

#                 # Collision check
#                 counter = 1
#                 while os.path.exists(new_path) and old_path != new_path:
#                     new_filename = f"{clean_date}_{counter}.pdf"
#                     new_path = os.path.join(data_folder, new_filename)
#                     counter += 1

#                 # Physically rename the file
#                 if old_path != new_path:
#                     os.rename(old_path, new_path)
#                     print(f"  -> Renamed: {filename}  ==>  {new_filename}")
#                     renamed_count += 1

#             except ValueError as e:
#                 print(f"  -> Failed to parse date for {filename}. Error: {e}")

#     print(f"Renamed {renamed_count} files.")


# # ==========================================
# # 1.2 Preprocessing – DOE PDF Extraction
# # ==========================================

# # ── Constants ──────────────────────────────────────────────────────────────

# BRANDS_ORDERED = [
#     'PETRON', 'SHELL', 'CALTEX', 'PHOENIX', 'TOTAL',
#     'FLYING V', 'UNIOIL', 'SEAOIL', 'PTT', 'INDEPENDENT'
# ]

# BRAND_DISPLAY = {
#     'PETRON': 'Petron', 'SHELL': 'Shell', 'CALTEX': 'Caltex',
#     'PHOENIX': 'Phoenix', 'TOTAL': 'Total', 'FLYING V': 'Flying V',
#     'UNIOIL': 'Unioil', 'SEAOIL': 'Seaoil', 'PTT': 'PTT',
#     'INDEPENDENT': 'Independent'
# }

# PRODUCT_MAP = {
#     'RON 100': 'RON 100', 'RON 97': 'RON 97', 'RON 95': 'RON 95',
#     'RON 91': 'RON 91', 'DIESEL PLUS': 'Diesel Plus',
#     'DIESEL': 'Diesel', 'KEROSENE': 'Kerosene'
# }

# CITIES = [
#     'Caloocan', 'Quezon', 'Manila', 'Pasig', 'Taguig', 'Makati',
#     'Paranaque', 'Parañaque', 'Muntinlupa', 'Las Pinas', 'Las Piñas',
#     'Marikina', 'Pasay', 'Valenzuela', 'Navotas', 'Malabon',
#     'Mandaluyong', 'San Juan', 'Pateros'
# ]

# COLS = [
#     'No.', 'Monitoring Dates', 'Effectivity Date', 'City', 'Product',
#     'Brand', 'Price Low (P/L)', 'Price High (P/L)', 'Notes'
# ]

# # ── XLSX style helpers ─────────────────────────────────────────────────────

# _DARK  = '1F3864'
# _BLUE  = '2E75B6'
# _LBLUE = 'D6E4F0'
# _ALT   = 'EBF3FB'
# _WHITE = 'FFFFFF'

# def _fill(h):
#     return PatternFill('solid', start_color=h, fgColor=h)

# def _bord(c='BDD7EE'):
#     s = Side(border_style='thin', color=c)
#     return Border(left=s, right=s, top=s, bottom=s)

# def _fnt(sz=10, bold=False, color='000000'):
#     return Font(name='Arial', size=sz, bold=bold, color=color)

# _CTR = Alignment(horizontal='center', vertical='center')
# _LFT = Alignment(horizontal='left',   vertical='center')
# _RGT = Alignment(horizontal='right',  vertical='center')

# # ── Date helpers ───────────────────────────────────────────────────────────

# def _derive_effectivity(s):
#     if not s:
#         return ''
#     for pat in [
#         r'([A-Za-z]+)\s+(\d+)\s*[-]\s*[A-Za-z]+\s+\d+,\s*(\d{4})',
#         r'([A-Za-z]+)\s+(\d+)[-\s].*?(\d{4})',
#         r'([A-Za-z]+)\s+(\d+),\s*(\d{4})',
#     ]:
#         m = re.search(pat, s)
#         if m:
#             try:
#                 return datetime.strptime(
#                     f"{m.group(1)[:3].capitalize()} {m.group(2)} {m.group(3)}",
#                     "%b %d %Y"
#                 ).strftime("%b %d, %Y")
#             except Exception:
#                 pass
#     return ''


# def _parse_dates(text):
#     monitoring = effectivity = ''

#     m = re.search(r'Date of Monitoring[:\s]+([A-Za-z][\w\s,.\-]+?\d{4})', text)
#     if m:
#         monitoring = re.sub(r'\s+', ' ', m.group(1)).strip()

#     if not monitoring:
#         m = re.search(r'[Ff]or the week of\s+(.+?\d{4})', text[:800])
#         if m:
#             monitoring = re.sub(r'\s+', ' ', m.group(1)).strip().rstrip(')')

#     m = re.search(r'Date of Effectivity\s*:\s*([A-Za-z][\w\s,]+?\d{4})', text[:1200])
#     if m:
#         ds = re.sub(r'\s+', ' ', m.group(1)).strip().replace('Januray', 'January')
#         try:
#             effectivity = datetime.strptime(ds, "%B %d, %Y").strftime("%b %d, %Y")
#         except Exception:
#             effectivity = ds

#     if not effectivity:
#         effectivity = _derive_effectivity(monitoring)

#     return monitoring, effectivity


# def _parse_sort_date(row):
#     for s in [row.get('Effectivity Date', ''), row.get('Monitoring Dates', '')]:
#         if not s or pd.isna(s) or str(s).strip() == '':
#             continue
#         s = str(s).strip().replace('Januray', 'January')
#         for fmt in ['%b %d, %Y', '%B %d, %Y']:
#             try:
#                 return datetime.strptime(s, fmt)
#             except Exception:
#                 pass
#         m = re.search(r'([A-Za-z]+)\s+(\d+).*?(\d{4})', s)
#         if m:
#             try:
#                 return datetime.strptime(
#                     f"{m.group(1)[:3].capitalize()} {m.group(2)} {m.group(3)}",
#                     "%b %d %Y"
#                 )
#             except Exception:
#                 pass
#     return datetime(1900, 1, 1)

# # ── City / brand helpers ───────────────────────────────────────────────────

# def _normalize_city(raw):
#     raw = raw.strip()
#     raw = raw.replace('Paranaque', 'Parañaque').replace('Las Pinas', 'Las Piñas')
#     raw = re.sub(r'\s*[Cc]ty\.?\s*$', ' City', raw)
#     if not re.search(r'[Cc]ity\s*$', raw):
#         raw += ' City'
#     return raw.strip()


# def _get_brand_positions(lines):
#     for line in lines:
#         if sum(1 for b in ['PETRON', 'SHELL', 'CALTEX'] if b in line.upper()) >= 2:
#             positions = {}
#             for brand in BRANDS_ORDERED:
#                 m = re.search(brand.replace(' ', r'\s+'), line, re.I)
#                 if m:
#                     positions[brand] = (m.start() + m.end()) / 2
#             m_range = re.search(r'OVERALL\s*RANGE', line, re.I)
#             range_start = m_range.start() if m_range else 9999
#             return positions, range_start
#     return {}, 9999


# def _extract_brand_prices(line, brand_centers, range_start):
#     nums = [
#         (m.start(), float(m.group()))
#         for m in re.finditer(r'\d{2,3}\.\d{2}', line)
#         if m.start() < range_start
#     ]
#     if not nums:
#         return {}

#     brands  = list(brand_centers.keys())
#     centers = list(brand_centers.values())
#     bp      = {b: [] for b in brands}

#     for pos, val in nums:
#         dists    = [abs(pos - c) for c in centers]
#         min_dist = min(dists)
#         if min_dist < 18:
#             nearest = brands[dists.index(min_dist)]
#             if len(bp[nearest]) < 2:
#                 bp[nearest].append(val)

#     return {b: v for b, v in bp.items() if v}

# # ── Core PDF extractor ─────────────────────────────────────────────────────

# def extract_pdf(pdf_path):
#     """
#     Extract brand-level price rows from a single DOE oil price PDF.

#     Parameters
#     ----------
#     pdf_path : str or Path

#     Returns
#     -------
#     list of dict
#     """
#     pages_text = []
#     with pdfplumber.open(str(pdf_path)) as pdf:
#         for page in pdf.pages:
#             t = page.extract_text(layout=True)
#             if t:
#                 pages_text.append(t)

#     text  = '\n'.join(pages_text)
#     lines = text.split('\n')

#     monitoring, effectivity = _parse_dates(text)
#     brand_centers, range_start = _get_brand_positions(lines)

#     city_pat = '|'.join(re.escape(c) for c in CITIES)
#     rows     = []
#     current_area = None

#     for line in lines:
#         cm = re.search(rf'^\s*({city_pat})\s*(City|Cty\.?)?', line, re.I)
#         if cm:
#             current_area = _normalize_city(cm.group(0).strip())

#         pm = re.search(
#             r'\b(RON\s*100|RON\s*97|RON\s*95|RON\s*91|DIESEL PLUS|DIESEL|KEROSENE)\b',
#             line, re.I
#         )
#         if not pm or not current_area or not brand_centers:
#             continue

#         raw_prod = re.sub(r'\s+', ' ', pm.group(1).strip()).upper()
#         product  = PRODUCT_MAP.get(raw_prod, raw_prod.title())

#         for brand_key, prices in _extract_brand_prices(line, brand_centers, range_start).items():
#             rows.append({
#                 'Monitoring Dates': monitoring,
#                 'Effectivity Date': effectivity,
#                 'City':             current_area,
#                 'Product':          product,
#                 'Brand':            BRAND_DISPLAY[brand_key],
#                 'Price Low (P/L)':  min(prices),
#                 'Price High (P/L)': max(prices),
#                 'Notes':            ''
#             })

#     return rows


# def extract_all_pdfs(pdf_dir):
#     """
#     Extract brand-level price rows from all DOE oil price PDFs in pdf_dir.

#     Parameters
#     ----------
#     pdf_dir : str or Path

#     Returns
#     -------
#     pd.DataFrame – serialized, sorted chronologically by effectivity date.
#     """
#     pdf_dir  = Path(pdf_dir)
#     all_rows = []
#     pdfs     = sorted(pdf_dir.glob("*.pdf"))

#     print(f"Found {len(pdfs)} PDF(s) in '{pdf_dir}'")

#     for i, pdf in enumerate(pdfs, 1):
#         if 'priceadj' in pdf.name.lower():
#             continue
#         try:
#             rows = extract_pdf(pdf)
#             all_rows.extend(rows)
#             print(f"  [{i}/{len(pdfs)}] {pdf.name} — {len(rows)} rows")
#         except Exception as e:
#             print(f"  [WARN] {pdf.name}: {e}")

#     if not all_rows:
#         print("\n[ERROR] No rows extracted. Check that pdf_dir points to the correct folder.")
#         return pd.DataFrame(columns=[c for c in COLS if c != 'No.'])

#     df = pd.DataFrame(all_rows)
#     df['_sort_date'] = df.apply(_parse_sort_date, axis=1)
#     df = (df.sort_values(['_sort_date', 'City', 'Product', 'Brand'])
#             .drop(columns=['_sort_date'])
#             .reset_index(drop=True))
#     df.insert(0, 'No.', range(1, len(df) + 1))

#     return df


# # ==========================================
# # 1.3 Export – CSV & XLSX
# # ==========================================

# def save_csv(df, output_path):
#     """Save DataFrame to CSV (UTF-8 with BOM for Excel compatibility)."""
#     df[COLS].to_csv(output_path, index=False, encoding='utf-8-sig')
#     print(f"CSV  saved → {output_path}  ({len(df):,} rows)")


# def save_xlsx(df, output_path):
#     """Save DataFrame to a formatted XLSX with a summary sheet."""
#     wb = Workbook()

#     # ── Sheet 1: City-Brand Detail ─────────────────────────────────────────
#     ws = wb.active
#     ws.title = 'City-Brand Detail'

#     ws.merge_cells('A1:I1')
#     ws['A1'] = 'DOE NCR Fuel Prices – Per City, Per Brand Detail'
#     ws['A1'].font      = _fnt(14, bold=True, color=_WHITE)
#     ws['A1'].fill      = _fill(_DARK)
#     ws['A1'].alignment = _CTR
#     ws.row_dimensions[1].height = 28

#     ws.merge_cells('A2:I2')
#     ws['A2'] = ('Source: Department of Energy, Oil Industry Management Bureau  |  '
#                 'Prices in P/L  |  Sorted chronologically by effectivity date')
#     ws['A2'].font      = Font(name='Arial', size=9, italic=True, color=_WHITE)
#     ws['A2'].fill      = _fill(_BLUE)
#     ws['A2'].alignment = _CTR
#     ws.row_dimensions[2].height = 16
#     ws.row_dimensions[3].height = 6

#     for ci, h in enumerate(COLS, 1):
#         c = ws.cell(row=4, column=ci, value=h)
#         c.font = _fnt(10, bold=True, color=_WHITE)
#         c.fill = _fill(_DARK); c.alignment = _CTR; c.border = _bord(_DARK)
#     ws.row_dimensions[4].height = 22

#     for ri, (_, row) in enumerate(df[COLS].iterrows(), 5):
#         fc = _ALT if ri % 2 == 0 else _WHITE
#         for ci, col in enumerate(COLS, 1):
#             val = row[col]
#             if pd.isna(val):
#                 val = None
#             c = ws.cell(row=ri, column=ci, value=val)
#             c.font = _fnt(9); c.fill = _fill(fc); c.border = _bord()
#             if ci == 1:
#                 c.alignment = _CTR
#             elif ci in (7, 8):
#                 c.alignment = _RGT
#                 if val is not None:
#                     c.number_format = '#,##0.00'
#             else:
#                 c.alignment = _LFT

#     for ci, w in enumerate([6, 20, 17, 18, 13, 13, 14, 15, 10], 1):
#         ws.column_dimensions[get_column_letter(ci)].width = w

#     ws.freeze_panes    = 'A5'
#     ws.auto_filter.ref = f'A4:I{4 + len(df)}'

#     # ── Sheet 2: Summary by Brand & Year ──────────────────────────────────
#     ws2 = wb.create_sheet('Summary by Brand & Year')
#     df2 = df.copy()
#     df2['_year'] = df2['Monitoring Dates'].str.extract(r'(\d{4})')[0].astype(float)
#     pivot    = (df2.groupby(['_year', 'Brand'])['Price Low (P/L)']
#                    .mean().unstack('Brand').sort_index())
#     brands_s = list(pivot.columns)
#     years_s  = list(pivot.index.astype(int))
#     ncols    = len(brands_s) + 1

#     ws2.merge_cells(f'A1:{get_column_letter(ncols)}1')
#     ws2['A1'] = 'Average Price Low (P/L) by Brand and Year – NCR'
#     ws2['A1'].font      = _fnt(13, bold=True, color=_WHITE)
#     ws2['A1'].fill      = _fill(_DARK)
#     ws2['A1'].alignment = _CTR
#     ws2.row_dimensions[1].height = 26

#     ws2.merge_cells(f'A2:{get_column_letter(ncols)}2')
#     ws2['A2'] = 'Source: DOE Oil Monitor Reports  |  Blank = no data for that brand/year'
#     ws2['A2'].font      = Font(name='Arial', size=9, italic=True, color=_WHITE)
#     ws2['A2'].fill      = _fill(_BLUE)
#     ws2['A2'].alignment = _CTR
#     ws2.row_dimensions[3].height = 6

#     hc = ws2.cell(row=4, column=1, value='Year')
#     hc.font = _fnt(10, bold=True, color=_WHITE); hc.fill = _fill(_DARK)
#     hc.alignment = _CTR; hc.border = _bord(_DARK)
#     ws2.row_dimensions[4].height = 22

#     for ci, b in enumerate(brands_s, 2):
#         c = ws2.cell(row=4, column=ci, value=b)
#         c.font = _fnt(10, bold=True, color=_WHITE); c.fill = _fill(_DARK)
#         c.alignment = _CTR; c.border = _bord(_DARK)

#     for ri, year in enumerate(years_s, 5):
#         fc = _ALT if ri % 2 == 0 else _WHITE
#         yc = ws2.cell(row=ri, column=1, value=int(year))
#         yc.font = _fnt(10, bold=True); yc.fill = _fill(_LBLUE)
#         yc.alignment = _CTR; yc.border = _bord()
#         for ci, b in enumerate(brands_s, 2):
#             val = pivot.loc[year, b] if b in pivot.columns and year in pivot.index else None
#             val = round(val, 2) if val is not None and pd.notna(val) else None
#             c = ws2.cell(row=ri, column=ci, value=val)
#             c.font = _fnt(9); c.fill = _fill(fc)
#             c.alignment = _RGT; c.border = _bord()
#             c.number_format = '#,##0.00'

#     ws2.column_dimensions['A'].width = 8
#     for ci in range(2, ncols + 1):
#         ws2.column_dimensions[get_column_letter(ci)].width = 13
#     ws2.freeze_panes = 'B5'

#     wb.save(output_path)
#     print(f"XLSX saved → {output_path}  ({len(df):,} rows)")

# # ==========================================
# # 2.0 Post-Processing & Feature Engineering (For Quantile Regression)
# # ==========================================

# def impute_missing_prices(df):
#     """
#     Creates a complete Cartesian grid of Dates x Cities x Brands x Products,
#     exposes missing values, and imputes them temporally (ffill) and spatially (NCR average).
#     """

#     df['Effectivity Date'] = pd.to_datetime(df['Effectivity Date'])

#     unique_dates = df['Effectivity Date'].unique()
#     unique_cities = df['City'].unique()
#     unique_brands = df['Brand'].unique()
#     unique_products = df['Product'].unique()

#     full_index = pd.MultiIndex.from_tuples(
#         itertools.product(unique_dates, unique_cities, unique_brands, unique_products),
#         names=['Effectivity Date', 'City', 'Brand', 'Product']
#     )
#     df_full = df.set_index(['Effectivity Date', 'City', 'Brand', 'Product']).reindex(full_index).reset_index()

#     df_full = df_full.sort_values(['City', 'Brand', 'Product', 'Effectivity Date'])

#     df_full['Price Low (P/L)'] = df_full.groupby(['City', 'Brand', 'Product'])['Price Low (P/L)'].ffill()
#     df_full['Price High (P/L)'] = df_full.groupby(['City', 'Brand', 'Product'])['Price High (P/L)'].ffill()

#     ncr_avg = df_full.groupby(['Effectivity Date', 'Brand', 'Product'])[['Price Low (P/L)', 'Price High (P/L)']].transform('mean')
#     df_full['Price Low (P/L)'] = df_full['Price Low (P/L)'].fillna(ncr_avg['Price Low (P/L)'])
#     df_full['Price High (P/L)'] = df_full['Price High (P/L)'].fillna(ncr_avg['Price High (P/L)'])


#     df_full = df_full.dropna(subset=['Price Low (P/L)', 'Price High (P/L)'])
    
#     print(f"Imputation complete. Total rows: {len(df_full):,}")
#     return df_full

# # do not mind this - this is for eda later on
# # def build_ncr_volatility_index(df_imputed):
# #     """
# #     Aggregates imputed city-level data into a single NCR-wide weekly index
# #     and calculates the week-over-week percentage change.
# #     """
# #     print("Building NCR volatility index...")
    
# #     # Aggregate to the NCR level
# #     df_ncr = df_imputed.groupby(['Effectivity Date', 'Product'])[['Price Low (P/L)', 'Price High (P/L)']].mean().reset_index()

# #     # Create 'Avg Price'
# #     df_ncr['Avg Price'] = (df_ncr['Price Low (P/L)'] + df_ncr['Price High (P/L)']) / 2

# #     # Sort chronologically
# #     df_ncr = df_ncr.sort_values(['Product', 'Effectivity Date'])

# #     # Calculate week-over-week percentage change
# #     df_ncr['Price_Pct_Change'] = df_ncr.groupby('Product')['Avg Price'].pct_change()
    
# #     # Drop the first week's NaN resulting from the pct_change calculation
# #     df_ncr = df_ncr.dropna(subset=['Price_Pct_Change']).reset_index(drop=True)
    
# #     return df_ncr

"""
utils.py
========
Requires: pdfplumber, pandas, openpyxl, requests, beautifulsoup4, python-dotenv
Install : pip install pdfplumber pandas openpyxl requests beautifulsoup4 python-dotenv
"""

import os
import re
import requests
import pdfplumber
import itertools
import pandas as pd
from sqlalchemy import create_engine
from pathlib import Path
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from dotenv import load_dotenv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load environment variables from the .env file
load_dotenv()

# ==========================================
# 00.1 Data Acquisition EIA Brent Prices
# ==========================================
def fetch_eia_prices(series_id, price_name):
    """Fetches daily spot prices for a given EIA series ID."""
    api_key = os.getenv('EIA_API_KEY')
    start_date = os.getenv('START_DATE')

    print(f"Fetching {price_name} prices from {start_date}...")

    url = (
        f"https://api.eia.gov/v2/petroleum/pri/spt/data/"
        f"?api_key={api_key}"
        f"&frequency=daily"
        f"&data[0]=value"
        f"&facets[series][]={series_id}"
        f"&start={start_date}"
        f"&sort[0][column]=period"
        f"&sort[0][direction]=desc"
        f"&length=5000"
    )

    response = requests.get(url)

    if response.status_code == 200:
        data = response.json()['response']['data']
        df = pd.DataFrame(data)

        # Clean and isolate the columns
        df = df[['period', 'value']]
        df.rename(columns={'period': 'Date', 'value': price_name}, inplace=True)

        # Convert to datetime and strip timezone
        df['Date'] = pd.to_datetime(df['Date'])
        if df['Date'].dt.tz is not None:
            df['Date'] = df['Date'].dt.tz_localize(None)

        return df
    else:
        print(f"Failed to fetch {price_name}. Status: {response.status_code}")
        return pd.DataFrame()


def download_eia_dataset(output_folder='data'):
    """Executes the pipeline to pull and merge Brent and WTI prices."""
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    df_brent = fetch_eia_prices('RBRTE', 'Brent_Price_USD')
    df_wti = fetch_eia_prices('RWTC', 'WTI_Price_USD')

    if not df_brent.empty and not df_wti.empty:
        print("Merging Brent and WTI datasets...")
        df_combined = pd.merge(df_brent, df_wti, on='Date', how='inner')
        df_combined.sort_values(by='Date', inplace=True)

        output_file = os.path.join(output_folder, 'Brent_WTI_Prices_2020_to_Present.csv')
        df_combined.to_csv(output_file, index=False)
        print(f"Success! Saved EIA data to {output_file}")
    else:
        print("Error: Could not complete merge due to missing data.")


# ==========================================
# 00.2 Data Acquisition DOE
# ==========================================
def download_doe_pdfs(output_folder='data'):
    """Scrapes the DOE portal for retail pump price PDFs."""
    target_url = os.getenv('DOE_TARGET_URL')
    session_cookie = os.getenv('DOE_SESSION_COOKIE')

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0 Safari/537.36',
        'Cookie': session_cookie
    }

    print("Authenticating with DOE portal...")
    response = requests.get(target_url, headers=headers)

    if "Log in" in response.text or response.status_code != 200:
        print("Recheck Cookie Validity. Authentication failed.")
        return

    soup = BeautifulSoup(response.text, 'html.parser')
    pdf_links = []

    for link in soup.find_all('a', href=True):
        href = link['href']
        if 'documents/d/' in href.lower() or '-pdf' in href.lower() or href.lower().endswith('.pdf'):
            full_url = urljoin(target_url, href)
            pdf_links.append(full_url)

    pdf_links = list(set(pdf_links))
    print(f"Found {len(pdf_links)} protected PDF(s). Downloading...")

    for pdf_url in pdf_links:
        file_name = pdf_url.split('/')[-1]

        if not file_name.endswith('.pdf'):
            file_name = file_name + '.pdf'

        file_path = os.path.join(output_folder, file_name)

        pdf_response = requests.get(pdf_url, headers=headers)
        pdf_response.raise_for_status()

        with open(file_path, 'wb') as file:
            file.write(pdf_response.content)
        print(f"Saved: {file_name}")

    print("DOE PDF extraction complete.")


# ==========================================
# 1.1 Preprocessing DOE filename
# ==========================================
def standardize_doe_date(data_folder='data/doe_pump_prices'):
    """
    Loops through the target folder, extracts dates from messy DOE filenames,
    and physically renames the PDF files to 'YYYY-MM-DD.pdf' on the OS.
    """
    renamed_count = 0

    for filename in os.listdir(data_folder):
        if filename.lower().endswith('.pdf'):

            year, month, day = None, None, None

            # Pattern 1: YYYY-MM-DD or YYYY-Month-DD (e.g., 2021-june-10, 2024-dec-24)
            m1 = re.search(r"(20\d{2})[-_]([a-z]{3,9}|\d{1,2})[-_](\d{1,2})", filename.lower())

            # Pattern 2: MM-DD-YYYY (e.g., 07-01-2025)
            m2 = re.search(r"(\d{1,2})[-_](\d{1,2})[-_](20\d{2})", filename.lower())

            # Pattern 3: MMDDYYYY contiguous (e.g., 08122025)
            m3 = re.search(r"(\d{2})(\d{2})(20\d{2})", filename.lower())

            # Check which pattern matched
            if m1:
                year, month, day = m1.groups()
            elif m2:
                month, day, year = m2.groups()
            elif m3:
                month, day, year = m3.groups()
            else:
                print(f"  -> WARNING: Could not find any date pattern in '{filename}'")
                continue

            day = day.zfill(2)
            month = str(month)

            try:
                # Handle textual months (forcing them to 3 letters: June -> Jun, Sept -> Sep)
                if month.isalpha():
                    month_short = month[:3].title()
                    clean_str = f"{year}-{month_short}-{day}"
                    clean_date = datetime.strptime(clean_str, "%Y-%b-%d").strftime("%Y-%m-%d")
                # Handle numeric months
                else:
                    month = month.zfill(2)
                    clean_str = f"{year}-{month}-{day}"
                    clean_date = datetime.strptime(clean_str, "%Y-%m-%d").strftime("%Y-%m-%d")

                # Construct the new file path
                new_filename = f"{clean_date}.pdf"
                old_path = os.path.join(data_folder, filename)
                new_path = os.path.join(data_folder, new_filename)

                # Collision check
                counter = 1
                while os.path.exists(new_path) and old_path != new_path:
                    new_filename = f"{clean_date}_{counter}.pdf"
                    new_path = os.path.join(data_folder, new_filename)
                    counter += 1

                # Physically rename the file
                if old_path != new_path:
                    os.rename(old_path, new_path)
                    print(f"  -> Renamed: {filename}  ==>  {new_filename}")
                    renamed_count += 1

            except ValueError as e:
                print(f"  -> Failed to parse date for {filename}. Error: {e}")

    print(f"Renamed {renamed_count} files.")


# ==========================================
# 1.2 Preprocessing – DOE PDF Extraction
# ==========================================

# ── Constants ──────────────────────────────────────────────────────────────

BRANDS_ORDERED = [
    'PETRON', 'SHELL', 'CALTEX', 'PHOENIX', 'TOTAL',
    'FLYING V', 'UNIOIL', 'SEAOIL', 'PTT', 'INDEPENDENT'
]

BRAND_DISPLAY = {
    'PETRON': 'Petron', 'SHELL': 'Shell', 'CALTEX': 'Caltex',
    'PHOENIX': 'Phoenix', 'TOTAL': 'Total', 'FLYING V': 'Flying V',
    'UNIOIL': 'Unioil', 'SEAOIL': 'Seaoil', 'PTT': 'PTT',
    'INDEPENDENT': 'Independent'
}

PRODUCT_MAP = {
    'RON 100': 'RON 100', 'RON 97': 'RON 97', 'RON 95': 'RON 95',
    'RON 91': 'RON 91', 'DIESEL PLUS': 'Diesel Plus',
    'DIESEL': 'Diesel', 'KEROSENE': 'Kerosene'
}

CITIES = [
    'Caloocan', 'Quezon', 'Manila', 'Pasig', 'Taguig', 'Makati',
    'Paranaque', 'Parañaque', 'Muntinlupa', 'Las Pinas', 'Las Piñas',
    'Marikina', 'Pasay', 'Valenzuela', 'Navotas', 'Malabon',
    'Mandaluyong', 'San Juan', 'Pateros'
]

COLS = [
    'No.', 'Monitoring Dates', 'Effectivity Date', 'City', 'Product',
    'Brand', 'Price Low (P/L)', 'Price High (P/L)', 'Notes'
]

# ── XLSX style helpers ─────────────────────────────────────────────────────

_DARK  = '1F3864'
_BLUE  = '2E75B6'
_LBLUE = 'D6E4F0'
_ALT   = 'EBF3FB'
_WHITE = 'FFFFFF'

def _fill(h):
    return PatternFill('solid', start_color=h, fgColor=h)

def _bord(c='BDD7EE'):
    s = Side(border_style='thin', color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def _fnt(sz=10, bold=False, color='000000'):
    return Font(name='Arial', size=sz, bold=bold, color=color)

_CTR = Alignment(horizontal='center', vertical='center')
_LFT = Alignment(horizontal='left',   vertical='center')
_RGT = Alignment(horizontal='right',  vertical='center')

# ── Date helpers ───────────────────────────────────────────────────────────

def _derive_effectivity(s):
    if not s:
        return ''
    for pat in [
        r'([A-Za-z]+)\s+(\d+)\s*[-]\s*[A-Za-z]+\s+\d+,\s*(\d{4})',
        r'([A-Za-z]+)\s+(\d+)[-\s].*?(\d{4})',
        r'([A-Za-z]+)\s+(\d+),\s*(\d{4})',
    ]:
        m = re.search(pat, s)
        if m:
            try:
                return datetime.strptime(
                    f"{m.group(1)[:3].capitalize()} {m.group(2)} {m.group(3)}",
                    "%b %d %Y"
                ).strftime("%b %d, %Y")
            except Exception:
                pass
    return ''


def _parse_dates(text):
    monitoring = effectivity = ''

    m = re.search(r'Date of Monitoring[:\s]+([A-Za-z][\w\s,.\-]+?\d{4})', text)
    if m:
        monitoring = re.sub(r'\s+', ' ', m.group(1)).strip()

    if not monitoring:
        m = re.search(r'[Ff]or the week of\s+(.+?\d{4})', text[:800])
        if m:
            monitoring = re.sub(r'\s+', ' ', m.group(1)).strip().rstrip(')')

    m = re.search(r'Date of Effectivity\s*:\s*([A-Za-z][\w\s,]+?\d{4})', text[:1200])
    if m:
        ds = re.sub(r'\s+', ' ', m.group(1)).strip().replace('Januray', 'January')
        try:
            effectivity = datetime.strptime(ds, "%B %d, %Y").strftime("%b %d, %Y")
        except Exception:
            effectivity = ds

    if not effectivity:
        effectivity = _derive_effectivity(monitoring)

    return monitoring, effectivity


def _parse_sort_date(row):
    for s in [row.get('Effectivity Date', ''), row.get('Monitoring Dates', '')]:
        if not s or pd.isna(s) or str(s).strip() == '':
            continue
        s = str(s).strip().replace('Januray', 'January')
        for fmt in ['%b %d, %Y', '%B %d, %Y']:
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                pass
        m = re.search(r'([A-Za-z]+)\s+(\d+).*?(\d{4})', s)
        if m:
            try:
                return datetime.strptime(
                    f"{m.group(1)[:3].capitalize()} {m.group(2)} {m.group(3)}",
                    "%b %d %Y"
                )
            except Exception:
                pass
    return datetime(1900, 1, 1)

# ── City / brand helpers ───────────────────────────────────────────────────

def _normalize_city(raw):
    raw = raw.strip()
    raw = raw.replace('Paranaque', 'Parañaque').replace('Las Pinas', 'Las Piñas')
    raw = re.sub(r'\s*[Cc]ty\.?\s*$', ' City', raw)
    if not re.search(r'[Cc]ity\s*$', raw):
        raw += ' City'
    return raw.strip()


def _get_brand_positions(lines):
    for line in lines:
        if sum(1 for b in ['PETRON', 'SHELL', 'CALTEX'] if b in line.upper()) >= 2:
            positions = {}
            for brand in BRANDS_ORDERED:
                m = re.search(brand.replace(' ', r'\s+'), line, re.I)
                if m:
                    positions[brand] = (m.start() + m.end()) / 2
            m_range = re.search(r'OVERALL\s*RANGE', line, re.I)
            range_start = m_range.start() if m_range else 9999
            return positions, range_start
    return {}, 9999


def _extract_brand_prices(line, brand_centers, range_start):
    nums = [
        (m.start(), float(m.group()))
        for m in re.finditer(r'\d{2,3}\.\d{2}', line)
        if m.start() < range_start
    ]
    if not nums:
        return {}

    brands  = list(brand_centers.keys())
    centers = list(brand_centers.values())
    bp      = {b: [] for b in brands}

    for pos, val in nums:
        dists    = [abs(pos - c) for c in centers]
        min_dist = min(dists)
        if min_dist < 18:
            nearest = brands[dists.index(min_dist)]
            if len(bp[nearest]) < 2:
                bp[nearest].append(val)

    return {b: v for b, v in bp.items() if v}

# ── Core PDF extractor ─────────────────────────────────────────────────────

def extract_pdf(pdf_path):
    """
    Extract brand-level price rows from a single DOE oil price PDF.

    Parameters
    ----------
    pdf_path : str or Path

    Returns
    -------
    list of dict
    """
    pages_text = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            t = page.extract_text(layout=True)
            if t:
                pages_text.append(t)

    text  = '\n'.join(pages_text)
    lines = text.split('\n')

    monitoring, effectivity = _parse_dates(text)
    brand_centers, range_start = _get_brand_positions(lines)

    city_pat = '|'.join(re.escape(c) for c in CITIES)
    rows     = []
    current_area = None

    for line in lines:
        cm = re.search(rf'^\s*({city_pat})\s*(City|Cty\.?)?', line, re.I)
        if cm:
            current_area = _normalize_city(cm.group(0).strip())

        pm = re.search(
            r'\b(RON\s*100|RON\s*97|RON\s*95|RON\s*91|DIESEL PLUS|DIESEL|KEROSENE)\b',
            line, re.I
        )
        if not pm or not current_area or not brand_centers:
            continue

        raw_prod = re.sub(r'\s+', ' ', pm.group(1).strip()).upper()
        product  = PRODUCT_MAP.get(raw_prod, raw_prod.title())

        for brand_key, prices in _extract_brand_prices(line, brand_centers, range_start).items():
            rows.append({
                'Monitoring Dates': monitoring,
                'Effectivity Date': effectivity,
                'City':             current_area,
                'Product':          product,
                'Brand':            BRAND_DISPLAY[brand_key],
                'Price Low (P/L)':  min(prices),
                'Price High (P/L)': max(prices),
                'Notes':            ''
            })

    return rows


def extract_all_pdfs(pdf_dir):
    """
    Extract brand-level price rows from all DOE oil price PDFs in pdf_dir.

    Parameters
    ----------
    pdf_dir : str or Path

    Returns
    -------
    pd.DataFrame – serialized, sorted chronologically by effectivity date.
    """
    pdf_dir  = Path(pdf_dir)
    all_rows = []
    pdfs     = sorted(pdf_dir.glob("*.pdf"))

    print(f"Found {len(pdfs)} PDF(s) in '{pdf_dir}'")

    for i, pdf in enumerate(pdfs, 1):
        if 'priceadj' in pdf.name.lower():
            continue
        try:
            rows = extract_pdf(pdf)
            all_rows.extend(rows)
            print(f"  [{i}/{len(pdfs)}] {pdf.name} — {len(rows)} rows")
        except Exception as e:
            print(f"  [WARN] {pdf.name}: {e}")

    if not all_rows:
        print("\n[ERROR] No rows extracted. Check that pdf_dir points to the correct folder.")
        return pd.DataFrame(columns=[c for c in COLS if c != 'No.'])

    df = pd.DataFrame(all_rows)
    df['_sort_date'] = df.apply(_parse_sort_date, axis=1)
    df = (df.sort_values(['_sort_date', 'City', 'Product', 'Brand'])
            .drop(columns=['_sort_date'])
            .reset_index(drop=True))
    df.insert(0, 'No.', range(1, len(df) + 1))

    return df


# ==========================================
# 1.3 Export – CSV & XLSX
# ==========================================

def save_csv(df, output_path):
    """Save DataFrame to CSV (UTF-8 with BOM for Excel compatibility)."""
    df[COLS].to_csv(output_path, index=False, encoding='utf-8-sig')
    print(f"CSV  saved → {output_path}  ({len(df):,} rows)")


def save_xlsx(df, output_path):
    """Save DataFrame to a formatted XLSX with a summary sheet."""
    wb = Workbook()

    # ── Sheet 1: City-Brand Detail ─────────────────────────────────────────
    ws = wb.active
    ws.title = 'City-Brand Detail'

    ws.merge_cells('A1:I1')
    ws['A1'] = 'DOE NCR Fuel Prices – Per City, Per Brand Detail'
    ws['A1'].font      = _fnt(14, bold=True, color=_WHITE)
    ws['A1'].fill      = _fill(_DARK)
    ws['A1'].alignment = _CTR
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:I2')
    ws['A2'] = ('Source: Department of Energy, Oil Industry Management Bureau  |  '
                'Prices in P/L  |  Sorted chronologically by effectivity date')
    ws['A2'].font      = Font(name='Arial', size=9, italic=True, color=_WHITE)
    ws['A2'].fill      = _fill(_BLUE)
    ws['A2'].alignment = _CTR
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    for ci, h in enumerate(COLS, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = _fnt(10, bold=True, color=_WHITE)
        c.fill = _fill(_DARK); c.alignment = _CTR; c.border = _bord(_DARK)
    ws.row_dimensions[4].height = 22

    for ri, (_, row) in enumerate(df[COLS].iterrows(), 5):
        fc = _ALT if ri % 2 == 0 else _WHITE
        for ci, col in enumerate(COLS, 1):
            val = row[col]
            if pd.isna(val):
                val = None
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = _fnt(9); c.fill = _fill(fc); c.border = _bord()
            if ci == 1:
                c.alignment = _CTR
            elif ci in (7, 8):
                c.alignment = _RGT
                if val is not None:
                    c.number_format = '#,##0.00'
            else:
                c.alignment = _LFT

    for ci, w in enumerate([6, 20, 17, 18, 13, 13, 14, 15, 10], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes    = 'A5'
    ws.auto_filter.ref = f'A4:I{4 + len(df)}'

    # ── Sheet 2: Summary by Brand & Year ──────────────────────────────────
    ws2 = wb.create_sheet('Summary by Brand & Year')
    df2 = df.copy()
    df2['_year'] = df2['Monitoring Dates'].str.extract(r'(\d{4})')[0].astype(float)
    pivot    = (df2.groupby(['_year', 'Brand'])['Price Low (P/L)']
                   .mean().unstack('Brand').sort_index())
    brands_s = list(pivot.columns)
    years_s  = list(pivot.index.astype(int))
    ncols    = len(brands_s) + 1

    ws2.merge_cells(f'A1:{get_column_letter(ncols)}1')
    ws2['A1'] = 'Average Price Low (P/L) by Brand and Year – NCR'
    ws2['A1'].font      = _fnt(13, bold=True, color=_WHITE)
    ws2['A1'].fill      = _fill(_DARK)
    ws2['A1'].alignment = _CTR
    ws2.row_dimensions[1].height = 26

    ws2.merge_cells(f'A2:{get_column_letter(ncols)}2')
    ws2['A2'] = 'Source: DOE Oil Monitor Reports  |  Blank = no data for that brand/year'
    ws2['A2'].font      = Font(name='Arial', size=9, italic=True, color=_WHITE)
    ws2['A2'].fill      = _fill(_BLUE)
    ws2['A2'].alignment = _CTR
    ws2.row_dimensions[3].height = 6

    hc = ws2.cell(row=4, column=1, value='Year')
    hc.font = _fnt(10, bold=True, color=_WHITE); hc.fill = _fill(_DARK)
    hc.alignment = _CTR; hc.border = _bord(_DARK)
    ws2.row_dimensions[4].height = 22

    for ci, b in enumerate(brands_s, 2):
        c = ws2.cell(row=4, column=ci, value=b)
        c.font = _fnt(10, bold=True, color=_WHITE); c.fill = _fill(_DARK)
        c.alignment = _CTR; c.border = _bord(_DARK)

    for ri, year in enumerate(years_s, 5):
        fc = _ALT if ri % 2 == 0 else _WHITE
        yc = ws2.cell(row=ri, column=1, value=int(year))
        yc.font = _fnt(10, bold=True); yc.fill = _fill(_LBLUE)
        yc.alignment = _CTR; yc.border = _bord()
        for ci, b in enumerate(brands_s, 2):
            val = pivot.loc[year, b] if b in pivot.columns and year in pivot.index else None
            val = round(val, 2) if val is not None and pd.notna(val) else None
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = _fnt(9); c.fill = _fill(fc)
            c.alignment = _RGT; c.border = _bord()
            c.number_format = '#,##0.00'

    ws2.column_dimensions['A'].width = 8
    for ci in range(2, ncols + 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 13
    ws2.freeze_panes = 'B5'

    wb.save(output_path)
    print(f"XLSX saved → {output_path}  ({len(df):,} rows)")

# ==========================================
# 2.0 Post-Processing & Feature Engineering (For Quantile Regression)
# ==========================================

def impute_missing_prices(df):
    """
    Creates a complete Cartesian grid of Dates x Cities x Brands x Products,
    exposes missing values, and imputes them temporally (ffill) and spatially (NCR average).
    """

    df['Effectivity Date'] = pd.to_datetime(df['Effectivity Date'])

    unique_dates = df['Effectivity Date'].unique()
    unique_cities = df['City'].unique()
    unique_brands = df['Brand'].unique()
    unique_products = df['Product'].unique()

    full_index = pd.MultiIndex.from_tuples(
        itertools.product(unique_dates, unique_cities, unique_brands, unique_products),
        names=['Effectivity Date', 'City', 'Brand', 'Product']
    )
    df_full = df.set_index(['Effectivity Date', 'City', 'Brand', 'Product']).reindex(full_index).reset_index()

    df_full = df_full.sort_values(['City', 'Brand', 'Product', 'Effectivity Date'])

    df_full['Price Low (P/L)'] = df_full.groupby(['City', 'Brand', 'Product'])['Price Low (P/L)'].ffill()
    df_full['Price High (P/L)'] = df_full.groupby(['City', 'Brand', 'Product'])['Price High (P/L)'].ffill()

    ncr_avg = df_full.groupby(['Effectivity Date', 'Brand', 'Product'])[['Price Low (P/L)', 'Price High (P/L)']].transform('mean')
    df_full['Price Low (P/L)'] = df_full['Price Low (P/L)'].fillna(ncr_avg['Price Low (P/L)'])
    df_full['Price High (P/L)'] = df_full['Price High (P/L)'].fillna(ncr_avg['Price High (P/L)'])


    df_full = df_full.dropna(subset=['Price Low (P/L)', 'Price High (P/L)'])
    
    print(f"Imputation complete. Total rows: {len(df_full):,}")
    return df_full

# do not mind this - this is for eda later on
# def build_ncr_volatility_index(df_imputed):
#     """
#     Aggregates imputed city-level data into a single NCR-wide weekly index
#     and calculates the week-over-week percentage change.
#     """
#     print("Building NCR volatility index...")
    
#     # Aggregate to the NCR level
#     df_ncr = df_imputed.groupby(['Effectivity Date', 'Product'])[['Price Low (P/L)', 'Price High (P/L)']].mean().reset_index()

#     # Create 'Avg Price'
#     df_ncr['Avg Price'] = (df_ncr['Price Low (P/L)'] + df_ncr['Price High (P/L)']) / 2

#     # Sort chronologically
#     df_ncr = df_ncr.sort_values(['Product', 'Effectivity Date'])

#     # Calculate week-over-week percentage change
#     df_ncr['Price_Pct_Change'] = df_ncr.groupby('Product')['Avg Price'].pct_change()
    
#     # Drop the first week's NaN resulting from the pct_change calculation
#     df_ncr = df_ncr.dropna(subset=['Price_Pct_Change']).reset_index(drop=True)
    
#     return df_ncr